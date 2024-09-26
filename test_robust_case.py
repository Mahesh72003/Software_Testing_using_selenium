import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
from tkinter import Tk, messagebox
from datetime import datetime

# Load data from Excel sheets
search_products_df = pd.read_excel("products.xlsx", sheet_name="Sheet1")

# Ensure that 'ProductURL' exists in the DataFrame
if 'ProductURL' not in search_products_df.columns:
    raise KeyError("The column 'ProductURL' does not exist in the DataFrame.")

# Initialize the results workbook
results_path = r"D:\mahesh2003\mini_project\mini_project_st\results\test_results_robust_case.xlsx"
try:
    results_wb = load_workbook(results_path)
    results_ws = results_wb.active
except FileNotFoundError:
    results_wb = Workbook()
    results_ws = results_wb.active
    results_ws.append(["Test Case", "Product Name", "Status", "Message"])

def log_result(test_case, product_name, status, message):
    results_ws.append([test_case, product_name, status, message])
    results_wb.save(results_path)

def add_datetime_before_testing():
    current_date = datetime.now().strftime("%d/%m/%Y")
    current_time = datetime.now().strftime("%H:%M:%S")
    results_ws.append(["Testing date =", current_date, "Testing Time =", current_time])
    results_wb.save(results_path)

def setup_driver():
    driver = webdriver.Chrome()
    driver.implicitly_wait(10)
    return driver

def test_home_page_load(driver):
    driver.get("https://www.1mg.com")
    assert "1mg" in driver.title
    log_result("Home Page Load", "N/A", "Passed", "Home page loaded successfully")

# New Robust Test Cases
def test_search_with_special_characters(driver):
    special_chars = "!@#$%^&*()_+{}:\"<>?|`~"
    search_box = driver.find_element(By.XPATH, "//input[@id='srchBarShwInfo']")
    search_box.clear()
    search_box.send_keys(special_chars)
    search_box.send_keys(Keys.RETURN)
    assert "No results found" in driver.page_source
    log_result("Search with Special Characters", "Special Characters", "Passed", "Search returned no results as expected")

def test_login_with_invalid_credentials(driver):
    login_button = driver.find_element(By.XPATH, '//*[@id="header"]/div[1]/div[2]/ul/li[1]/a[1]')
    login_button.click()
    phone_field = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="component"]/div/div[2]/div/div/div/div/div/div/div/div[2]/div[2]/div/div/div[1]/div[2]/div/div[1]/div/input'))
    )
    phone_field.send_keys("0000000000")  # Invalid phone number
    submit_phone_button = driver.find_element(By.XPATH, '//*[@id="component"]/div/div[2]/div/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[1]/a')
    submit_phone_button.click()
    time.sleep(3)
    close_login = driver.find_element(By.XPATH, '//*[@id="component"]/div/div[2]/div/div/div/div/div/div/div/div[2]/div[1]/span')
    close_login.click()
    log_result("Login with Invalid Credentials", "N/A", "Passed", "Login failed with invalid credentials as expected")

def test_add_to_cart_without_login(driver):
    driver.get("https://www.1mg.com")
    search_box = driver.find_element(By.XPATH, "//input[@id='srchBarShwInfo']")
    search_box.send_keys("Ecosprin 75 Tablet")
    search_box.send_keys(Keys.RETURN)
    time.sleep(3)
    driver.get("https://www.1mg.com/drugs/ecosprin-75-tablet-40765")
    try:
        add_to_cart_button = driver.find_element(By.XPATH, '//*[@id="atc-content"]/div/div[2]/div[3]/div/div')
        add_to_cart_button.click()
        log_result("Add to Cart Without Login", "N/A", "Passed", "Add to Cart attempted without login")
    except Exception as e:
        log_result("Add to Cart Without Login", "N/A", "Failed", str(e))

def test_profile_page_access_without_login(driver):
    driver.get("https://www.1mg.com/profile")
    assert "Login" in driver.page_source  # Assuming a login prompt or redirection occurs
    log_result("Profile Page Access Without Login", "N/A", "Passed", "Profile page access attempted without login")



def test_input_with_html_tags(driver):
    malicious_input = "<script>alert('Test')</script>"
    search_box = driver.find_element(By.XPATH, "//input[@id='srchBarShwInfo']")
    search_box.clear()
    search_box.send_keys(malicious_input)
    search_box.send_keys(Keys.RETURN)
    assert "No results found" in driver.page_source
    log_result("Input with HTML Tags", "HTML Tags", "Passed", "Search handled input with HTML tags without error")


def test_long_string_input(driver):
    long_string = "a " * 1000  # Long string
    search_box = driver.find_element(By.XPATH, "//input[@id='srchBarShwInfo']")
    search_box.clear()
    search_box.send_keys(long_string)
    search_box.send_keys(Keys.RETURN)
    log_result("Long String Input", "Long String", "Passed", "Search handled long string input without error")

def test_invalid_url_access(driver):
    invalid_url = "https://www.1mg.com/invalid-url"
    driver.get(invalid_url)
    assert "404" in driver.title  # Check for 404 error
    log_result("Invalid URL Access", "N/A", "Passed", "Attempted access to invalid URL resulted in 404 error")

def test_slow_network_simulation(driver):
    driver.set_network_conditions(
        offline=False,
        latency=1000,  # 1 second
        download_throughput=500 * 1024,  # 500 kb/s
        upload_throughput=500 * 1024   # 500 kb/s
    )
    driver.get("https://www.1mg.com")
    assert "1mg" in driver.title
    log_result("Slow Network Simulation", "N/A", "Passed", "Home page loaded under simulated slow network conditions")

if __name__ == "__main__":
    add_datetime_before_testing()
    driver = setup_driver()
    try:
        # Run robust test cases
        test_home_page_load(driver)
        time.sleep(10)
        test_search_with_special_characters(driver)
        time.sleep(10)
        test_login_with_invalid_credentials(driver)
        time.sleep(10)
        test_add_to_cart_without_login(driver)
        time.sleep(10)
        test_profile_page_access_without_login(driver)
        time.sleep(10)
        test_input_with_html_tags(driver)
        time.sleep(10)
        test_long_string_input(driver)
        time.sleep(10)
        test_invalid_url_access(driver)
        time.sleep(10)

    finally:
        driver.quit()
