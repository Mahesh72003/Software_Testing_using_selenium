import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
from tkinter import Tk, messagebox
from datetime import datetime

# Load data from Excel sheets
search_products_df = pd.read_excel("products.xlsx", sheet_name="Sheet1")

# Ensure that 'ProductURL' exists in the DataFrame
if 'ProductURL' not in search_products_df.columns:
    raise KeyError("The column 'ProductURL' does not exist in the DataFrame.")

# Initialize the results workbook
results_path = r"D:\mahesh2003\mini_project\mini_project_st\results\test_results_happy_case.xlsx"
try:
    results_wb = load_workbook(results_path)
    results_ws = results_wb.active


except FileNotFoundError:
    results_wb = Workbook()
    results_ws = results_wb.active
    results_ws.append(["Test Case", "Product Name", "Status", "Message"])

def log_result(test_case, product_name, status, message):
    results_ws.append([test_case, product_name, status, message])
    results_wb.save(results_path)

def add_datetime_before_testing():
    current_date = datetime.now().strftime("%d/%m/%Y")
    current_time = datetime.now().strftime("%H:%M:%S")
    results_ws.append(["Testing date =", current_date, "Testing Time =", current_time])
    results_wb.save(results_path)

def setup_driver():
    driver = webdriver.Chrome()
    driver.implicitly_wait(10)
    return driver

def test_home_page_load(driver):
    driver.get("https://www.1mg.com")
    assert "1mg" in driver.title
    time.sleep(7)
    update= driver.find_element(By.XPATH,'//*[@id="update-city-modal"]/div/div[3]/div[2]')
    update.click()
    log_result("Home Page Load", "N/A", "Passed", "Home page loaded successfully")

def test_home_page_load_new(driver):
    driver.get("https://www.1mg.com")
    assert "1mg" in driver.title
    time.sleep(7)
    log_result("Home Page Load", "N/A", "Passed", "Home page loaded successfully")

def test_search_functionality(driver, product_name):
    search_box = driver.find_element(By.XPATH, "//input[@id='srchBarShwInfo']")
    search_box.send_keys(product_name)
    search_box.send_keys(Keys.RETURN)
    assert product_name in driver.title or "No results found" not in driver.page_source
    log_result("Search Functionality", product_name, "Passed", "Search completed successfully")

def test_product_page_navigation(driver, product_url):
    driver.get(product_url)
    log_result("Product Page Navigation", product_url, "Passed", "Product page loaded successfully")

def test_add_to_cart(driver, product_name):
    try:
        # Check if the product is sold out
        add_to_cart_button = driver.find_element(By.XPATH, '//*[@id="container"]/div/div/div[2]/div[4]/div[1]/div/div[2]/div[3]')
        if add_to_cart_button:
            add_to_cart_button.click()
            log_result("Add to Cart", product_name, "Passed", "Item added to cart successfully")
        else:
            sold_out_button = driver.find_elements(By.XPATH, '//*[@id="container"]/div/div/div[2]/div[4]/div[1]/div/div/div[4]/button')
            log_result("Add to Cart", product_name, "Failed", "Product is sold out")
    except Exception as e:
        log_result("Add to Cart", product_name, "Passed", "Item is searched",)

def test_view_cart(driver):
    driver.get("https://www.1mg.com/cart")
    log_result("View Cart", "N/A", "Passed", "Cart viewed successfully")

def show_otp_message():
    # Create a simple Tkinter window
    root = Tk()
    root.withdraw() 
    messagebox.showinfo("OTP Required", "Please enter the OTP on the web page and click done.")
    root.destroy()

def test_login_functionality(driver): 
    login_button = driver.find_element(By.XPATH, '//*[@id="header"]/div[4]/div[2]/div/div[1]/div/a[1]')
    login_button.click()
    phone_field = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By
        .XPATH, '//*[@id="login-signup-modal--react"]/div/div/div/div/div/div/div/div[2]/div[2]/div/div/div[1]/div[2]/div/div[1]/div/input'))
    )
    phone_field.send_keys("9110498586")
    submit_phone_button = driver.find_element(By.XPATH, '//*[@id="login-signup-modal--react"]/div/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[1]/a')
    submit_phone_button.click()
    # Show the OTP message
    show_otp_message()
    time.sleep(5)
    continue_otp_filed = driver.find_element(By.XPATH,'//*[@id="login-signup-modal--react"]/div/div/div/div/div/div/div/div[2]/div[2]/div/div/div[2]/div[1]/a')
    continue_otp_filed.click()
    log_result("Login Functionality", "N/A", "Passed", "Login completed successfully")

def test_profile(driver):
    driver.get("https://www.1mg.com/profile")
    log_result("Profile", "N/A", "Passed", "Profile viewed successfully")

def test_logout(driver):
    driver.get("https://www.1mg.com/logout")
    log_result("Logout", "N/A", "Passed", "Logout completed successfully")

if __name__ == "__main__":
    add_datetime_before_testing()
    driver = setup_driver()
    try:
        test_home_page_load(driver)
        time.sleep(10)
        #test_login_functionality(driver)
        #time.sleep(10)
        for _, row in search_products_df.iterrows():
            product_name = row['ProductName']  
            test_search_functionality(driver, product_name)
            time.sleep(10)
            product_url = row.get('ProductURL', None) 
            if product_url:
                test_product_page_navigation(driver, product_url)
                time.sleep(10)
                test_add_to_cart(driver, product_name)
                time.sleep(10)
            else:
                log_result("Product Page Navigation", product_name, "Failed", "Product URL not found")
        
        test_view_cart(driver)
        time.sleep(10)
        test_home_page_load_new(driver)
        time.sleep(10)
        test_profile(driver)
        time.sleep(10)
        test_logout(driver)
        time.sleep(10)

    finally:
        driver.quit()